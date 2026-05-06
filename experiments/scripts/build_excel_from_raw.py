"""
Combine raw CSV outputs into one Excel file with the required submission columns.

Usage:
    python scripts/build_excel_from_raw.py \
        --raw-dir data/raw \
        --output output/app_update_history_filled.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_COLUMNS = [
    "app_name",
    "platform",
    "developer_company",
    "app_category",
    "version_number",
    "version_release_date",
    "is_current_version",
    "initial_app_release_date",
    "update_description_release_notes",
    "standardized_update_categories",
    "standardized_update_summary",
    "source_url",
    "data_quality_notes",
]

SUMMARY_TEXT = [
    ["Section", "Content"],
    ["Project goal", "Collect app-platform-version update history observations for 10 matched iOS/Android popular apps."],
    ["Collection approach", "Raw rows were collected using scripted methods where possible: Apple iTunes Lookup API for current iOS metadata, app-store-scraper for iOS version history when available, and google-play-scraper for current Android metadata. Android historical coverage may require supplemental public sources because Google Play often exposes only current release notes."],
    ["Categorization method", "Release notes were mapped to standardized update categories using transparent keyword-based rules in scripts/categorize_updates.py, then summarized into analysis-ready descriptions."],
    ["Known limitations", "Some app stores provide generic release notes, and Android historical update records are less consistently available from Google Play public metadata. Missing or weakly supported observations should be flagged in data_quality_notes."],
    ["Next manual step", "Review rows, supplement Android history from AppBrain/APKMirror where feasible, and add a short time-series pattern summary before submission."],
]


def read_raw_csvs(raw_dir: Path) -> pd.DataFrame:
    csv_files = sorted(raw_dir.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {raw_dir}")

    frames = []
    for csv_file in csv_files:
        df = pd.read_csv(csv_file)
        for col in OUTPUT_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        frames.append(df[OUTPUT_COLUMNS])

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(
        subset=["app_name", "platform", "version_number", "version_release_date"],
        keep="first",
    )
    return combined


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_length + 2, 12), 55)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Excel workbook from raw app update CSV files.")
    parser.add_argument("--raw-dir", default="data/raw", help="Directory containing raw CSV files.")
    parser.add_argument("--output", default="output/app_update_history_filled.xlsx", help="Output Excel path.")
    args = parser.parse_args()

    raw_dir = Path(args.raw_dir)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    history_df = read_raw_csvs(raw_dir)
    summary_df = pd.DataFrame(SUMMARY_TEXT[1:], columns=SUMMARY_TEXT[0])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        history_df.to_excel(writer, sheet_name="Update History", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    style_workbook(output_path)
    print(f"Saved combined Excel workbook to {output_path}")


if __name__ == "__main__":
    main()
