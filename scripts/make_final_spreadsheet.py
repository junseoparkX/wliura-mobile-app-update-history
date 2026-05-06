from pathlib import Path
from shutil import copyfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ------------------------------------------------------------
# 1. Paths
# ------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent

DATA_DIR = PROJECT_ROOT / "data"
PROCESSED_DIR = DATA_DIR / "processed"
ANDROID_DIR = PROCESSED_DIR / "android"
IOS_DIR = PROCESSED_DIR / "ios"
FINAL_DIR = DATA_DIR / "final"

SUMMARY_XLSX = PROCESSED_DIR / "summary_sheet_with_github_styled.xlsx"

IOS_CSV = IOS_DIR / "ios_apptopia_all_apps_combined.csv"
ANDROID_CSV = ANDROID_DIR / "android_apkpure_all_apps_combined.csv"

OUTPUT_XLSX = FINAL_DIR / "wliura_mobile_app_update_history_final.xlsx"

FINAL_DIR.mkdir(parents=True, exist_ok=True)


# ------------------------------------------------------------
# 2. Styling helper
# ------------------------------------------------------------

def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    ws.freeze_panes = "A2"

    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 45)

    ws.row_dimensions[1].height = 24


def add_dataframe_sheet(wb, sheet_name, df):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    style_sheet(ws)


# ------------------------------------------------------------
# 3. Read data
# ------------------------------------------------------------

ios_df = pd.read_csv(IOS_CSV)
android_df = pd.read_csv(ANDROID_CSV)

ios_df["source_file"] = IOS_CSV.name
android_df["source_file"] = ANDROID_CSV.name

combined_df = pd.concat([ios_df, android_df], ignore_index=True)

print(f"[INFO] iOS rows: {len(ios_df)}")
print(f"[INFO] Android rows: {len(android_df)}")
print(f"[INFO] Combined rows: {len(combined_df)}")


# ------------------------------------------------------------
# 4. Start from styled summary workbook
# ------------------------------------------------------------

if SUMMARY_XLSX.exists():
    copyfile(SUMMARY_XLSX, OUTPUT_XLSX)
    wb = load_workbook(OUTPUT_XLSX)
    print("[INFO] Summary workbook copied first.")
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Summary file was not found."
    print("[WARNING] Summary workbook not found. Created blank Summary sheet.")


# ------------------------------------------------------------
# 5. Add sheets in required order
# ------------------------------------------------------------

add_dataframe_sheet(wb, "iOS_Update_History", ios_df)
add_dataframe_sheet(wb, "Android_Update_History", android_df)
add_dataframe_sheet(wb, "All_Update_History", combined_df)


# ------------------------------------------------------------
# 6. Save
# ------------------------------------------------------------

wb.save(OUTPUT_XLSX)

print("[DONE] Final spreadsheet created:")
print(OUTPUT_XLSX)
print("[INFO] Sheet order:")
for sheet in wb.sheetnames:
    print(" -", sheet)